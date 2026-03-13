"""
Server-side PDF generation using Playwright.
Generates HTML that matches the current Cutting Plan PDF design.
"""

import asyncio
import sys
from playwright.sync_api import sync_playwright
from pathlib import Path
from typing import Dict, List, Any, Optional
import base64
import math
from concurrent.futures import ThreadPoolExecutor

class CuttingPlanPDFGenerator:
    """Generates Cutting Plan PDFs server-side using Playwright."""
    
    def __init__(self):
        self.page_width = 297  # A4 landscape width in mm
        self.page_height = 210  # A4 landscape height in mm
        
    def generate_pdf(
        self,
        nesting_report: Dict[str, Any],
        project_name: str,
        tolerance: float,
        tolerance_enabled: bool,
        trim: float,
        kerf: float,
        selected_profiles: List[str],
        icons: Dict[str, str],  # base64 encoded icons
        stockbar_svg_data: List[Dict[str, Any]] = None,  # Extracted SVG polygon data from browser
        total_weight: float = 0,  # Total weight in tonnes, pre-calculated from frontend
        company_details: Dict[str, str] = None  # Company details for footer
    ) -> bytes:
        """Generate PDF from nesting report data."""
        
        if company_details is None:
            company_details = {}
        
        with sync_playwright() as p:
            browser = p.chromium.launch()
            page = browser.new_page()
            
            # Generate HTML content
            html_content = self._generate_html(
                nesting_report=nesting_report,
                project_name=project_name,
                tolerance=tolerance,
                tolerance_enabled=tolerance_enabled,
                trim=trim,
                kerf=kerf,
                selected_profiles=selected_profiles,
                icons=icons,
                stockbar_svg_data=stockbar_svg_data,
                total_weight=total_weight,
                company_details=company_details
            )
            
            # Set content and wait for rendering
            page.set_content(html_content, wait_until='networkidle')
            
            # Generate PDF with footer (except cover page)
            from datetime import datetime
            current_date = datetime.now().strftime("%d %b %Y")
            company_name = company_details.get('companyName', 'N/A')

            footer_template = f"""
            <div style="width: 100%; font-size: 9px; padding: 9px 40px; border-top: 1px solid #E5E7EB; display: flex; justify-content: space-between; align-items: center; color: #6B7280; background: white;">
                <div style="display: flex; align-items: center;">
                    <img src="data:image/svg+xml;base64,{icons.get('logo_small', '')}" style="width: 80px; height: 28px;" />
                </div>
                <div style="display: flex; align-items: center; gap: 12px; line-height: 1;">
                    <span style="line-height: 1;"><strong>Company Name:</strong> {company_name}</span>
                    <span style="color: #D1D5DB; line-height: 1;">•</span>
                    <span style="line-height: 1;"><strong>Project name:</strong> {project_name}</span>
                    <span style="color: #D1D5DB; line-height: 1;">•</span>
                    <span style="line-height: 1;"><strong>Page:</strong> <span class="pageNumber"></span> of <span class="totalPages"></span></span>
                </div>
            </div>
            """
            
            pdf_bytes = page.pdf(
                format='A4',
                landscape=True,
                print_background=True,
                display_header_footer=True,
                header_template='<div></div>',
                footer_template=footer_template,
                margin={
                    'top': '0mm',
                    'right': '0mm',
                    'bottom': '12mm',
                    'left': '0mm'
                },
                prefer_css_page_size=True
            )
            
            browser.close()
            return pdf_bytes
    
    def _generate_html(
        self,
        nesting_report: Dict[str, Any],
        project_name: str,
        tolerance: float,
        tolerance_enabled: bool,
        trim: float,
        kerf: float,
        selected_profiles: List[str],
        icons: Dict[str, str],
        stockbar_svg_data: List[Dict[str, Any]] = None,
        total_weight: float = 0,
        company_details: Dict[str, str] = None
    ) -> str:
        """Generate HTML content that matches the React PDF design."""
        
        if company_details is None:
            company_details = {}
        
        from datetime import datetime
        current_date = datetime.now().strftime("%d %b %Y")
        
        # Calculate totals for cover page
        totals = self._calculate_totals(nesting_report, selected_profiles, total_weight)
        
        # Calculate total pages early for cover page
        profiles = [p for p in nesting_report.get('profiles', []) 
                   if p['profile_name'] in selected_profiles]
        total_stockbars = sum(len(p.get('cutting_patterns', [])) for p in profiles)
        total_pages = 1 + total_stockbars  # 1 cover + all stockbars
        
        # Generate cover page
        cover_page_html = self._generate_cover_page(
            project_name=project_name,
            totals=totals,
            total_pages=total_pages,
            tolerance=tolerance,
            tolerance_enabled=tolerance_enabled,
            trim=trim,
            kerf=kerf,
            icons=icons
        )
        
        # Generate cutting plan pages - group stockbars by profile
        cutting_pages_html = ""
        page_num = 2  # Start from 2 (cover is page 1)
        
        for profile_idx, profile in enumerate(profiles):
            profile_name = profile.get('profile_name', 'Unknown')
            cutting_patterns = profile.get('cutting_patterns', [])
            
            # Start new page for each profile
            stockbars_html = ""
            for idx, pattern in enumerate(cutting_patterns):
                # Find matching SVG data for this stockbar
                svg_data = None
                if stockbar_svg_data:
                    for svg_item in stockbar_svg_data:
                        if svg_item.get('profileName') == profile_name and svg_item.get('patternIdx') == idx:
                            svg_data = svg_item.get('svgData')
                            break
                
                stockbars_html += self._generate_stockbar_section(
                    pattern=pattern,
                    pattern_idx=idx,
                    profile_name=profile_name,
                    tolerance=tolerance,
                    tolerance_enabled=tolerance_enabled,
                    trim=trim,
                    kerf=kerf,
                    icons=icons,
                    svg_data=svg_data
                )
            
            cutting_pages_html += self._generate_profile_page(
                profile_name=profile_name,
                stockbars_html=stockbars_html,
                project_name=project_name,
                page_num=page_num,
                total_pages=total_pages,
                icons=icons
            )
            page_num += 1
        
        # Combine everything
        html = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        @page {{
            size: A4 landscape;
            margin: 40px 0 12mm 0;
        }}
        
        @page :first {{
            margin: 0 0 12mm 0;
        }}
        
        body {{
            font-family: 'Helvetica', 'Arial', sans-serif;
            -webkit-print-color-adjust: exact;
            print-color-adjust: exact;
            counter-reset: page-counter;
        }}
        
        .page {{
            width: {self.page_width}mm;
            height: {self.page_height}mm;
            position: relative;
            background: white;
        }}
        
        .profile-section {{
            width: {self.page_width}mm;
            position: relative;
            background: white;
        }}
        
        .page + .profile-section {{
            page-break-before: avoid;
        }}
        
        .profile-section ~ .profile-section {{
            page-break-before: always;
        }}
        
        /* Footer styles (for cover page only) */
        .footer {{
            position: absolute;
            bottom: 0;
            left: 0;
            right: 0;
            padding: 8px 40px;
            border-top: 1px solid #E5E7EB;
            display: flex;
            justify-content: space-between;
            align-items: center;
            font-size: 9px;
            color: #6B7280;
        }}
        
        .footer-left {{
            display: flex;
            align-items: center;
        }}
        
        .footer-right {{
            display: flex;
            align-items: center;
            gap: 12px;
        }}
        
        .footer-dot {{
            color: #6B7280;
            margin: 0 6px;
        }}
        
        /* Cover page styles */
        .cover-logo-container {{
            padding: 20px 0;
            background-color: #F5F5F5;
            text-align: center;
            margin-bottom: 20px;
        }}
        
        .cover-main-content {{
            padding-left: calc((100% - 280px) / 2);
            max-width: 600px;
            padding-bottom: 40px;
        }}
        
        .cover-info-row {{
            display: flex;
            align-items: center;
            margin-bottom: 6px;
            gap: 10px;
        }}
        
        .cover-icon-large {{
            width: 20px;
            height: 20px;
            flex-shrink: 0;
        }}
        
        .cover-icon-settings {{
            width: 20px;
            height: 20px;
            flex-shrink: 0;
            filter: grayscale(100%) brightness(0.5);
        }}
        
        .cover-label {{
            font-size: 12px;
            color: #000;
            font-weight: 700;
            white-space: nowrap;
            min-width: 130px;
        }}
        
        .cover-value {{
            font-size: 12px;
            color: #000;
            font-weight: 400;
            white-space: nowrap;
        }}
        
        .cover-unit {{
            font-size: 10px;
            color: #6B7280;
            font-weight: 400;
            margin-left: 4px;
        }}
        
        .cover-divider {{
            height: 1px;
            background: #E5E7EB;
            margin: 14px 0;
            width: 150%;
            max-width: 600px;
        }}
        
        .cover-settings-title {{
            font-size: 14px;
            font-weight: 700;
            color: #000;
            margin-bottom: 14px;
        }}
        
        /* Cutting plan page styles */
        .cutting-content {{
            padding: 0 40px 20px 40px;
        }}
        
        .profile-title {{
            font-size: 18px;
            font-weight: 600;
            margin-bottom: 20px;
            margin-top: 0;
            color: #000;
            page-break-after: avoid;
        }}
        
        .stockbars-container {{
            margin-top: 0;
        }}
        
        .stockbar-section {{
            margin-top: 0;
            margin-bottom: 20px;
            padding: 10px;
            background: white;
            border-radius: 10px;
            page-break-inside: avoid;
            break-inside: avoid;
        }}
        
        .stockbar-section:not(:first-child) {{
            margin-top: 20px;
        }}
        
        .stockbar-title {{
            font-size: 10px;
            font-weight: 500;
            color: #6B7280;
            margin-bottom: 6px;
        }}
        
        .stockbar-info {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 8px;
        }}
        
        .info-boxes {{
            display: flex;
            align-items: center;
            gap: 8px;
            padding: 3px 8px;
            border: 1px solid #E5E7EB;
            border-radius: 5px;
            background: rgba(250, 250, 250, 0.2);
        }}
        
        .info-box {{
            display: inline-flex;
            align-items: center;
            gap: 4px;
            font-size: 10px;
            font-weight: 500;
        }}
        
        .info-icon {{
            width: 11px;
            height: 11px;
        }}
        
        .info-text {{
            color: #374151;
        }}
        
        .divider {{
            width: 1px;
            height: 11px;
            background: #E5E7EB;
        }}
        
        .waste-box {{
            padding: 3px 8px;
            border: 1px solid #E5E7EB;
            border-radius: 5px;
            background: rgba(250, 250, 250, 0.2);
            font-size: 10px;
            font-weight: 500;
            display: flex;
            align-items: center;
            gap: 4px;
        }}
        
        .waste-box.good {{
            background: rgba(28, 185, 126, 0.12);
            border-color: rgba(0, 129, 122, 0.4);
        }}
        
        .waste-box.good .info-text {{
            color: #00312F;
        }}
        
        .stockbar-visual {{
            margin: 16px 0 12px 0;
            height: 60px;
            border: 1px solid #D1D5DB;
            border-radius: 4px;
            position: relative;
            background: white;
            overflow: hidden;
        }}
        
        .cutting-table {{
            width: 65%;
            margin-top: 8px;
            border-collapse: collapse;
            font-size: 10px;
        }}
        
        .cutting-table thead tr {{
            background: #F3F4F6;
            border-bottom: 1px solid #D1D5DB;
        }}
        
        .cutting-table th {{
            height: 26px;
            padding: 5px 8px;
            text-align: left;
            font-weight: 500;
            font-size: 10px;
            color: #4B5563;
            vertical-align: middle;
            white-space: nowrap;
        }}
        
        .cutting-table tbody tr {{
            border-bottom: 1px solid #E5E7EB;
        }}
        
        .cutting-table tbody tr:last-child {{
            border-bottom: none;
        }}
        
        .cutting-table td {{
            padding: 6px 8px;
            vertical-align: middle;
            color: #374151;
            font-size: 10px;
        }}
    </style>
</head>
<body>
{cover_page_html}
{cutting_pages_html}
</body>
</html>
"""
        return html
    
    def _calculate_totals(self, nesting_report: Dict[str, Any], selected_profiles: List[str], total_weight: float = 0) -> Dict[str, Any]:
        """Calculate totals for cover page."""
        profiles = [p for p in nesting_report.get('profiles', []) 
                   if p['profile_name'] in selected_profiles]
        
        # Use the weight passed from frontend (already in tonnes)
        profile_types = len(profiles)
        total_cuts = sum(
            sum(max(0, len(pattern.get('parts', [])) - 1) 
                for pattern in p.get('cutting_patterns', []))
            for p in profiles
        )
        
        return {
            'weight': total_weight,  # Already in tonnes from frontend
            'profile_types': profile_types,
            'cuts': total_cuts
        }
    
    def _generate_cover_page(
        self,
        project_name: str,
        totals: Dict[str, Any],
        total_pages: int,
        tolerance: float,
        tolerance_enabled: bool,
        trim: float,
        kerf: float,
        icons: Dict[str, str]
    ) -> str:
        """Generate cover page HTML."""
        
        # Get current date
        from datetime import datetime
        current_date = datetime.now().strftime("%d %b %Y")
        
        tolerance_row = f"""
            <div class="cover-info-row">
                <img src="data:image/svg+xml;base64,{icons.get('tolerance', '')}" class="cover-icon-settings" />
                <span class="cover-label">Stockbar Tolerance:</span>
                <span class="cover-value">{tolerance:.0f} <span class="cover-unit">(mm)</span></span>
            </div>
        """ if tolerance_enabled else ""
        
        return f"""
<div class="page">
    <div class="cover-logo-container">
        <img src="data:image/svg+xml;base64,{icons.get('logo_main', '')}" style="width: 280px; height: 98px;" />
    </div>
    
    <div class="cover-main-content">
        <div class="cover-info-row">
            <img src="data:image/svg+xml;base64,{icons.get('pdf_project_name', '')}" class="cover-icon-large" />
            <span class="cover-label">Project Name:</span>
            <span class="cover-value">{project_name}</span>
        </div>
        
        <div class="cover-info-row">
            <img src="data:image/svg+xml;base64,{icons.get('pdf_date', '')}" class="cover-icon-large" />
            <span class="cover-label">Date:</span>
            <span class="cover-value">{current_date}</span>
        </div>
        
        <div class="cover-info-row">
            <img src="data:image/svg+xml;base64,{icons.get('pdf_weight', '')}" class="cover-icon-large" />
            <span class="cover-label">Weight:</span>
            <span class="cover-value">{totals['weight']:.3f} <span class="cover-unit">(t)</span></span>
        </div>
        
        <div class="cover-info-row">
            <img src="data:image/svg+xml;base64,{icons.get('pdf_profile_type', '')}" class="cover-icon-large" />
            <span class="cover-label">Profile Types:</span>
            <span class="cover-value">{totals['profile_types']}</span>
        </div>
        
        <div class="cover-info-row">
            <img src="data:image/svg+xml;base64,{icons.get('pdf_cutting_qty', '')}" class="cover-icon-large" />
            <span class="cover-label">Cutting Quantity:</span>
            <span class="cover-value">{totals['cuts']}</span>
        </div>
        
        <div class="cover-divider"></div>
        
        {tolerance_row}
        
        <div class="cover-info-row">
            <img src="data:image/svg+xml;base64,{icons.get('trim', '')}" class="cover-icon-settings" />
            <span class="cover-label">Manual Trim:</span>
            <span class="cover-value">{trim:.0f} <span class="cover-unit">(mm)</span></span>
        </div>
        
        <div class="cover-info-row">
            <img src="data:image/svg+xml;base64,{icons.get('kerf', '')}" class="cover-icon-settings" />
            <span class="cover-label">Saw Kerf:</span>
            <span class="cover-value">{kerf:.0f} <span class="cover-unit">(mm)</span></span>
        </div>
    </div>
</div>
"""
    
    def _generate_profile_page(
        self,
        profile_name: str,
        stockbars_html: str,
        project_name: str,
        page_num: int,
        total_pages: int,
        icons: Dict[str, str]
    ) -> str:
        """Generate a page section for a profile with all its stockbars."""
        
        return f"""
<div class="profile-section">
    <div class="cutting-content">
        <h2 class="profile-title">{profile_name}</h2>
        <div class="stockbars-container">
            {stockbars_html}
        </div>
    </div>
</div>
"""
    
    def _generate_stockbar_section(
        self,
        pattern: Dict[str, Any],
        pattern_idx: int,
        profile_name: str,
        tolerance: float,
        tolerance_enabled: bool,
        trim: float,
        kerf: float,
        icons: Dict[str, str],
        svg_data: Dict[str, Any] = None
    ) -> str:
        """Generate HTML for a single stockbar section."""
        
        stock_length = pattern.get('stock_length', 0)
        waste = pattern.get('waste', 0)
        waste_percentage = pattern.get('waste_percentage', 0)
        parts = pattern.get('parts', [])
        
        # Tolerance row
        tolerance_html = ""
        if tolerance_enabled:
            tolerance_html = f"""
                <div class="info-box">
                    <img src="data:image/svg+xml;base64,{icons.get('tolerance_section', '')}" class="info-icon" />
                    <span class="info-text">{tolerance:.0f}mm</span>
                </div>
                <div class="divider"></div>
            """
        
        # Waste box styling
        waste_class = "waste-box good" if waste_percentage <= 20 else "waste-box"
        
        # Generate SVG visualization using extracted data from browser
        svg_html = self._generate_stockbar_svg(
            pattern=pattern,
            stock_length=stock_length,
            parts=parts,
            svg_data=svg_data
        )
        
        # Generate cutting table
        table_html = self._generate_cutting_table(
            parts=parts,
            profile_name=profile_name
        )
        
        return f"""
<div class="stockbar-section">
    <div class="stockbar-title">Stockbar {pattern_idx + 1}</div>
    
    <div class="stockbar-info">
        <div class="info-boxes">
            <div class="info-box">
                <img src="data:image/svg+xml;base64,{icons.get('length', '')}" class="info-icon" />
                <span class="info-text">{stock_length:,.0f}mm</span>
            </div>
            <div class="divider"></div>
            {tolerance_html}
            <div class="info-box">
                <img src="data:image/svg+xml;base64,{icons.get('trim_section', icons.get('trim', ''))}" class="info-icon" />
                <span class="info-text">{trim:.0f}mm</span>
            </div>
            <div class="divider"></div>
            <div class="info-box">
                <img src="data:image/svg+xml;base64,{icons.get('kerf_section', icons.get('kerf', ''))}" class="info-icon" />
                <span class="info-text">{kerf:.0f}mm</span>
            </div>
        </div>
        
        <div class="{waste_class}">
            <img src="data:image/svg+xml;base64,{icons.get('waste', '')}" class="info-icon" />
            <span class="info-text">{waste:,.0f}mm ({waste_percentage:.2f}%)</span>
        </div>
    </div>
    
    <div class="stockbar-visual">
        {svg_html}
    </div>
    
    {table_html}
</div>
"""
    
    def _generate_stockbar_svg(
        self,
        pattern: Dict[str, Any],
        stock_length: float,
        parts: List[Dict[str, Any]],
        svg_data: Dict[str, Any] = None
    ) -> str:
        """Generate SVG visualization of stockbar with parts using extracted browser SVG data."""
        
        # If we have extracted SVG data from browser, use it directly!
        if svg_data and svg_data.get('parts'):
            view_box = svg_data.get('viewBox', '0 0 1000 60')
            svg_parts = []
            
            for part_data in svg_data['parts']:
                points = part_data.get('points', '')
                fill = part_data.get('fill', '#ccc')
                part_name = part_data.get('partName', '')
                
                # Render polygon exactly as it was in the browser
                # Match the app's polygon styling: rgba(156, 163, 175, 0.1) fill, #9ca3af stroke
                svg_parts.append(f'<polygon points="{points}" fill="rgba(156, 163, 175, 0.1)" stroke="#9ca3af" stroke-width="1" stroke-linejoin="miter" shape-rendering="crispEdges" />')
                
                # Add part number label (matching app's styling)
                if part_name and points:
                    # Calculate center of polygon for label placement
                    try:
                        coords = [float(x) for pair in points.split() for x in pair.split(',')]
                        if len(coords) >= 2:
                            # Average x coordinates for center
                            x_coords = [coords[i] for i in range(0, len(coords), 2)]
                            y_coords = [coords[i] for i in range(1, len(coords), 2)]
                            center_x = sum(x_coords) / len(x_coords)
                            center_y = sum(y_coords) / len(y_coords)
                            # Display part NUMBER (1, 2, 3...) not the part name
                            # Use app's text styling: 12px, medium weight, #374151 color
                            svg_parts.append(f'<text x="{center_x}" y="{center_y + 4}" text-anchor="middle" font-size="12" font-weight="500" fill="#374151" font-family="system-ui, -apple-system, sans-serif">{part_name}</text>')
                    except:
                        pass
            
            svg_content = ''.join(svg_parts)
            return f'<svg width="100%" height="100%" viewBox="{view_box}" preserveAspectRatio="none">{svg_content}</svg>'
        
        # Fallback: if no SVG data provided, show error
        return '<svg width="100%" height="100%" viewBox="0 0 1000 60"><text x="500" y="30" text-anchor="middle" fill="#666">No SVG data available</text></svg>'
    
    def _generate_cutting_table(
        self,
        parts: List[Dict[str, Any]],
        profile_name: str
    ) -> str:
        """Generate cutting list table."""
        
        # Group parts by name
        part_groups = {}
        for part in parts:
            part_data = part.get('part', {})
            part_name = part_data.get('reference') or part_data.get('element_name') or 'Unknown'
            part_length = part.get('length', 0)
            start_angle = part_data.get('start_angle')
            end_angle = part_data.get('end_angle')
            
            if part_name in part_groups:
                part_groups[part_name]['count'] += 1
            else:
                part_groups[part_name] = {
                    'name': part_name,
                    'length': part_length,
                    'count': 1,
                    'start_angle': start_angle,
                    'end_angle': end_angle
                }
        
        # Sort by length descending
        sorted_groups = sorted(part_groups.values(), key=lambda x: x['length'], reverse=True)
        
        # Generate table rows
        rows_html = ""
        for idx, group in enumerate(sorted_groups):
            start_angle_str = self._format_angle(group['start_angle'])
            end_angle_str = self._format_angle(group['end_angle'])
            
            rows_html += f"""
                <tr>
                    <td style="width: 6%;">{idx + 1}</td>
                    <td style="width: 24%;">{profile_name}</td>
                    <td style="width: 14%;">{group['name']}</td>
                    <td style="width: 14%;">{group['length']:,.0f}</td>
                    <td style="width: 14%;">{group['count']}</td>
                    <td style="width: 14%;">{start_angle_str}</td>
                    <td style="width: 14%;">{end_angle_str}</td>
                </tr>
            """
        
        return f"""
<table class="cutting-table">
    <thead>
        <tr>
            <th style="width: 6%;">#</th>
            <th style="width: 24%;">Profile Name</th>
            <th style="width: 14%;">Part Name</th>
            <th style="width: 14%;">Length (mm)</th>
            <th style="width: 14%;">Quantity</th>
            <th style="width: 14%;">Start Angle</th>
            <th style="width: 14%;">End Angle</th>
        </tr>
    </thead>
    <tbody>
        {rows_html}
    </tbody>
</table>
"""
    
    def _format_angle(self, angle: Any) -> str:
        """Format angle for display."""
        if angle is None or angle == '':
            return '90.0°'
        
        try:
            if isinstance(angle, (int, float)):
                return f'{float(angle):.1f}°'
            elif isinstance(angle, str):
                # Extract number from string
                import re
                match = re.search(r'-?\d+(?:\.\d+)?', angle)
                if match:
                    return f'{float(match.group()):.1f}°'
        except:
            pass
        
        return '90.0°'


class BOMPDFGenerator:
    """Generates Bill of Materials PDFs server-side using Playwright."""
    
    def __init__(self):
        self.page_width = 210  # A4 portrait width in mm
        self.page_height = 297  # A4 portrait height in mm
        
    def generate_pdf(
        self,
        nesting_report: Dict[str, Any],
        report: Dict[str, Any],
        project_name: str,
        company_details: Dict[str, str],
        icons: Dict[str, str]  # base64 encoded icons
    ) -> bytes:
        """Generate BOM PDF from nesting report data."""
        
        with sync_playwright() as p:
            browser = p.chromium.launch()
            page = browser.new_page()
            
            # Generate HTML content
            html_content = self._generate_html(
                nesting_report=nesting_report,
                report=report,
                project_name=project_name,
                company_details=company_details,
                icons=icons
            )
            
            # Set content and wait for rendering
            page.set_content(html_content, wait_until='networkidle')
            
            # Generate PDF (portrait A4) with footer
            company_name = company_details.get('companyName', 'N/A')
            
            footer_template = f"""
            <div style="width: 100%; font-size: 9px; padding: 9px 62px; border-top: 1px solid #E5E7EB; display: flex; justify-content: space-between; align-items: center; color: #6B7280; background: white;">
                <div style="display: flex; align-items: center;">
                    <img src="data:image/svg+xml;base64,{icons.get('logo', '')}" style="width: 80px; height: 28px;" />
                </div>
                <div style="display: flex; align-items: center; gap: 12px; line-height: 1;">
                    <span style="line-height: 1;"><strong>Company Name:</strong> {company_name}</span>
                    <span style="color: #D1D5DB; line-height: 1;">•</span>
                    <span style="line-height: 1;"><strong>Project name:</strong> {project_name}</span>
                    <span style="color: #D1D5DB; line-height: 1;">•</span>
                    <span style="line-height: 1;"><strong>Page:</strong> <span class="pageNumber"></span> of <span class="totalPages"></span></span>
                </div>
            </div>
            """
            
            pdf_bytes = page.pdf(
                format='A4',
                landscape=False,
                print_background=True,
                display_header_footer=True,
                header_template='<div></div>',
                footer_template=footer_template,
                margin={
                    'top': '0mm',
                    'right': '0mm',
                    'bottom': '12mm',
                    'left': '0mm'
                },
                prefer_css_page_size=True
            )
            
            browser.close()
            return pdf_bytes
    
    def _generate_html(
        self,
        nesting_report: Dict[str, Any],
        report: Dict[str, Any],
        project_name: str,
        company_details: Dict[str, str],
        icons: Dict[str, str]
    ) -> str:
        """Generate complete HTML for BOM PDF."""
        
        from datetime import datetime
        current_date = datetime.now().strftime("%d %B, %Y")  # Used in company info section
        
        # Calculate totals based on stock lengths used (same as frontend)
        total_weight_t = 0
        total_length_m = 0
        profile_count = len(nesting_report.get('profiles', []))
        
        # Build table rows data - one row per (profile, stock_length) combination
        bom_rows = []
        for profile in nesting_report.get('profiles', []):
            profile_name = profile.get('profile_name', '')
            
            # Get weight per meter from report
            profile_data = None
            for p in report.get('profiles', []):
                if p.get('profile_name') == profile_name:
                    profile_data = p
                    break
            
            weight_per_meter = 0
            if profile_data and profile.get('total_length', 0) > 0:
                total_length_mm = profile.get('total_length', 0)
                total_length_m_calc = total_length_mm / 1000.0
                weight_per_meter = profile_data.get('total_weight', 0) / total_length_m_calc
            
            # Split by stock_lengths_used - create one row per stock length
            stock_lengths_used = profile.get('stock_lengths_used', {})
            for stock_length_str, bar_count in stock_lengths_used.items():
                if bar_count > 0:
                    stock_length = float(stock_length_str)
                    stock_length_m = stock_length / 1000.0
                    
                    # Calculate weight for this stock length
                    stock_weight_kg = (weight_per_meter * stock_length_m * bar_count)
                    
                    # Add to grand totals
                    total_length_m += stock_length_m * bar_count
                    total_weight_t += stock_weight_kg / 1000.0
                    
                    bom_rows.append({
                        'profile_name': profile_name,
                        'stock_length_m': stock_length_m,
                        'quantity': bar_count,
                        'weight_kg': stock_weight_kg
                    })
        
        # Generate table rows HTML
        table_rows_html = ''
        for row in bom_rows:
            table_rows_html += f'''
            <tr>
                <td>{row['profile_name']}</td>
                <td>{row['stock_length_m']:.2f}</td>
                <td>{row['quantity']}</td>
                <td>{row['weight_kg']:.0f}</td>
            </tr>
            '''
        
        # Build company info rows conditionally
        company_info_rows = f'''
                    <div class="company-row">
                        <img src="data:image/svg+xml;base64,{icons.get('company', '')}" class="company-icon" />
                        <span class="company-label">Company name:</span>
                        <span class="company-value">{company_details.get('companyName', 'Your Company Name')}</span>
                    </div>'''
        
        # Add address row only if address exists
        if company_details.get('address'):
            company_info_rows += f'''
                    <div class="company-row">
                        <img src="data:image/svg+xml;base64,{icons.get('address', '')}" class="company-icon" />
                        <span class="company-label">Address:</span>
                        <span class="company-value">{company_details.get('address', '')}</span>
                    </div>'''
        
        # Add email row
        company_info_rows += f'''
                    <div class="company-row">
                        <img src="data:image/svg+xml;base64,{icons.get('email', '')}" class="company-icon" />
                        <span class="company-label">Email:</span>
                        <span class="company-value">{company_details.get('email', 'Company Email')}</span>
                    </div>'''
        
        # Add phone row only if phone exists
        if company_details.get('phoneNumber'):
            company_info_rows += f'''
                    <div class="company-row">
                        <img src="data:image/svg+xml;base64,{icons.get('phone', '')}" class="company-icon" />
                        <span class="company-label">Phone:</span>
                        <span class="company-value">{company_details.get('phoneNumber', '')}</span>
                    </div>'''
        
        # Add date row
        company_info_rows += f'''
                    <div class="company-row">
                        <img src="data:image/svg+xml;base64,{icons.get('date', '')}" class="company-icon" />
                        <span class="company-label">Date:</span>
                        <span class="company-value">{current_date}</span>
                    </div>'''
        
        # Build HTML
        html = f'''
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                @page {{
                    size: A4 portrait;
                    margin: 50px 0 80px 0;
                }}
                
                * {{
                    margin: 0;
                    padding: 0;
                    box-sizing: border-box;
                }}
                
                body {{
                    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
                    padding: 0 62px;
                    background: white;
                }}
                
                /* Header Section */
                .header {{
                    margin-bottom: 24px;
                }}
                
                .title {{
                    font-size: 32px;
                    font-weight: 600;
                    color: #000000;
                    letter-spacing: -0.5px;
                    margin-bottom: 20px;
                }}
                
                .company-info {{
                    margin-top: 12px;
                }}
                
                .company-row {{
                    display: flex;
                    align-items: center;
                    margin-bottom: 8px;
                }}
                
                .company-icon {{
                    width: 14px;
                    height: 14px;
                    margin-right: 10px;
                    flex-shrink: 0;
                    filter: grayscale(100%) brightness(0.4);
                }}
                
                .company-label {{
                    font-size: 11px;
                    color: #000000;
                    font-weight: 700;
                    margin-right: 4px;
                }}
                
                .company-value {{
                    font-size: 11px;
                    color: #000000;
                    font-weight: 400;
                }}
                
                /* Divider */
                .divider {{
                    border-bottom: 1px solid #D3D3D3;
                    margin: 20px 0;
                }}
                
                /* Project Section */
                .project-section {{
                    display: flex;
                    align-items: center;
                    margin-bottom: 16px;
                }}
                
                .project-icon {{
                    width: 16px;
                    height: 16px;
                    margin-right: 8px;
                }}
                
                .project-name {{
                    font-size: 18px;
                    font-weight: 400;
                    color: #000000;
                }}
                
                /* Summary Box */
                .summary-box {{
                    background-color: rgba(28, 185, 126, 0.12);
                    border-radius: 4px;
                    padding: 16px;
                    margin-bottom: 20px;
                    display: inline-block;
                    min-width: 220px;
                }}
                
                .summary-row {{
                    display: flex;
                    justify-content: space-between;
                    margin-bottom: 4px;
                }}
                
                .summary-row:last-child {{
                    margin-bottom: 0;
                }}
                
                .summary-label {{
                    font-size: 10px;
                    font-weight: 700;
                    color: #000000;
                }}
                
                .summary-value {{
                    font-size: 10px;
                    color: #000000;
                    font-weight: 400;
                }}
                
                /* Table */
                .bom-table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin-top: 9px;
                    page-break-inside: auto;
                }}
                
                .bom-table tr {{
                    page-break-inside: avoid;
                    page-break-after: auto;
                }}
                
                .bom-table thead {{
                    display: table-header-group;
                }}
                
                .bom-table thead {{
                    border-bottom: 1px solid #D3D3D3;
                    background-color: #F3F4F6;
                }}
                
                .bom-table th {{
                    text-align: left;
                    padding: 6px 10px;
                    font-size: 8px;
                    font-weight: 700;
                    color: #A5A7A9;
                    text-transform: uppercase;
                }}
                
                .bom-table td {{
                    padding: 6px 10px;
                    font-size: 9px;
                    color: #000000;
                    border-bottom: 0.5px solid #F0F0F0;
                }}
                
                .bom-table th:nth-child(2),
                .bom-table th:nth-child(3),
                .bom-table th:nth-child(4),
                .bom-table td:nth-child(2),
                .bom-table td:nth-child(3),
                .bom-table td:nth-child(4) {{
                    text-align: right;
                }}
                
            </style>
        </head>
        <body>
            <!-- Header with Company Info -->
            <div class="header">
                <div class="title">Bill of materials</div>
                
                <div class="company-info">
{company_info_rows}
                </div>
            </div>
            
            <div class="divider"></div>
            
            <!-- Project Section -->
            <div class="project-section">
                <img src="data:image/svg+xml;base64,{icons.get('project', '')}" class="project-icon" />
                <div class="project-name">{project_name}</div>
            </div>
            
            <!-- Summary Box -->
            <div class="summary-box">
                <div class="summary-row">
                    <span class="summary-label">Total weight</span>
                    <span class="summary-value">{total_weight_t:.3f} (t)</span>
                </div>
                <div class="summary-row">
                    <span class="summary-label">Total length</span>
                    <span class="summary-value">{total_length_m:.0f} (m)</span>
                </div>
                <div class="summary-row">
                    <span class="summary-label">Profile Type</span>
                    <span class="summary-value">{profile_count}</span>
                </div>
            </div>
            
            <!-- BOM Table -->
            <table class="bom-table">
                <thead>
                    <tr>
                        <th>PROFILE NAME</th>
                        <th>STOCK LENGTH (M)</th>
                        <th>QUANTITY</th>
                        <th>WEIGHT (KG)</th>
                    </tr>
                </thead>
                <tbody>
                    {table_rows_html}
                </tbody>
            </table>
        </body>
        </html>
        '''
        
        return html


class PartsListPDFGenerator:
    """Generates Parts Selection List PDFs server-side using Playwright."""

    def __init__(self):
        self.page_width = 210  # A4 portrait width in mm
        self.page_height = 297  # A4 portrait height in mm

    def generate_pdf(
        self,
        parts_by_profile: Dict[str, List[Dict]],
        selected_parts: Dict[str, List[str]],
        project_name: str,
        company_details: Dict[str, str],
        icons: Dict[str, str]
    ) -> bytes:
        """Generate Parts List PDF from parts data."""

        with sync_playwright() as p:
            browser = p.chromium.launch()
            page = browser.new_page()

            html_content = self._generate_html(
                parts_by_profile=parts_by_profile,
                selected_parts=selected_parts,
                project_name=project_name,
                company_details=company_details,
                icons=icons
            )

            page.set_content(html_content, wait_until='networkidle')

            company_name = company_details.get('companyName', 'N/A')

            footer_template = f"""
            <div style="width: 100%; font-size: 9px; padding: 9px 62px; border-top: 1px solid #E5E7EB; display: flex; justify-content: space-between; align-items: center; color: #6B7280; background: white;">
                <div style="display: flex; align-items: center;">
                    <img src="data:image/svg+xml;base64,{icons.get('logo', '')}" style="width: 80px; height: 28px;" />
                </div>
                <div style="display: flex; align-items: center; gap: 12px; line-height: 1;">
                    <span style="line-height: 1;"><strong>Company Name:</strong> {company_name}</span>
                    <span style="color: #D1D5DB; line-height: 1;">•</span>
                    <span style="line-height: 1;"><strong>Project name:</strong> {project_name}</span>
                    <span style="color: #D1D5DB; line-height: 1;">•</span>
                    <span style="line-height: 1;"><strong>Page:</strong> <span class="pageNumber"></span> of <span class="totalPages"></span></span>
                </div>
            </div>
            """

            pdf_bytes = page.pdf(
                format='A4',
                landscape=False,
                print_background=True,
                display_header_footer=True,
                header_template='<div></div>',
                footer_template=footer_template,
                margin={
                    'top': '0mm',
                    'right': '0mm',
                    'bottom': '12mm',
                    'left': '0mm'
                },
                prefer_css_page_size=True
            )

            browser.close()
            return pdf_bytes

    def _generate_html(
        self,
        parts_by_profile: Dict[str, List[Dict]],
        selected_parts: Dict[str, List[str]],
        project_name: str,
        company_details: Dict[str, str],
        icons: Dict[str, str]
    ) -> str:
        """Generate complete HTML for Parts List PDF."""

        from datetime import datetime
        current_date = datetime.now().strftime("%d %B, %Y")

        # Calculate totals
        total_parts = 0
        total_weight_kg = 0.0

        # Build company info rows
        company_info_rows = f'''
                    <div class="company-row">
                        <img src="data:image/svg+xml;base64,{icons.get('company', '')}" class="company-icon" />
                        <span class="company-label">Company name:</span>
                        <span class="company-value">{company_details.get('companyName', 'Your Company Name')}</span>
                    </div>'''

        if company_details.get('address'):
            company_info_rows += f'''
                    <div class="company-row">
                        <img src="data:image/svg+xml;base64,{icons.get('address', '')}" class="company-icon" />
                        <span class="company-label">Address:</span>
                        <span class="company-value">{company_details.get('address', '')}</span>
                    </div>'''

        if company_details.get('email'):
            company_info_rows += f'''
                    <div class="company-row">
                        <img src="data:image/svg+xml;base64,{icons.get('email', '')}" class="company-icon" />
                        <span class="company-label">Email:</span>
                        <span class="company-value">{company_details.get('email', '')}</span>
                    </div>'''

        if company_details.get('phoneNumber'):
            company_info_rows += f'''
                    <div class="company-row">
                        <img src="data:image/svg+xml;base64,{icons.get('phone', '')}" class="company-icon" />
                        <span class="company-label">Phone:</span>
                        <span class="company-value">{company_details.get('phoneNumber', '')}</span>
                    </div>'''

        company_info_rows += f'''
                    <div class="company-row">
                        <img src="data:image/svg+xml;base64,{icons.get('date', '')}" class="company-icon" />
                        <span class="company-label">Date:</span>
                        <span class="company-value">{current_date}</span>
                    </div>'''

        # Generate profile sections
        profile_sections_html = ''
        
        for profile_name, parts_list in parts_by_profile.items():
            selected_part_numbers = set(selected_parts.get(profile_name, []))
            
            # Filter only selected parts
            selected_parts_list = [
                part for part in parts_list 
                if str(part.get('part_number', '')) in selected_part_numbers or 
                   str(part.get('product_id', '')) in selected_part_numbers
            ]
            
            if not selected_parts_list:
                continue
            
            # Sort by part number
            selected_parts_list.sort(key=lambda p: str(p.get('part_number', '')))
            
            # Calculate profile totals
            profile_total_weight = sum(
                part.get('weight', 0) * part.get('quantity', 1) 
                for part in selected_parts_list
            )
            total_parts += len(selected_parts_list)
            total_weight_kg += profile_total_weight
            
            # Generate table rows for this profile
            table_rows_html = ''
            for part in selected_parts_list:
                part_weight = part.get('weight', 0)
                quantity = part.get('quantity', 1)
                part_total_weight = part_weight * quantity
                
                table_rows_html += f'''
                <tr>
                    <td>{part.get('part_number', 'Unknown')}</td>
                    <td>{part.get('element_name', 'Unnamed')}</td>
                    <td>{part.get('length', 0):.1f}</td>
                    <td>{quantity}</td>
                    <td>{part_weight:.2f}</td>
                    <td>{part_total_weight:.2f}</td>
                </tr>
                '''
            
            profile_sections_html += f'''
            <div class="profile-section">
                <h2 class="profile-title">{profile_name} • {len(selected_parts_list)} parts</h2>
                
                <table class="parts-table">
                    <thead>
                        <tr>
                            <th>PART NUMBER</th>
                            <th>PART NAME</th>
                            <th>LENGTH (MM)</th>
                            <th>QUANTITY</th>
                            <th>WEIGHT (KG)</th>
                            <th>TOTAL WEIGHT (KG)</th>
                        </tr>
                    </thead>
                    <tbody>
                        {table_rows_html}
                    </tbody>
                </table>
            </div>
            '''

        # Build HTML with same styling as BOM
        html = f'''
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                @page {{
                    size: A4 portrait;
                    margin: 50px 0 80px 0;
                }}
                
                * {{
                    margin: 0;
                    padding: 0;
                    box-sizing: border-box;
                }}
                
                body {{
                    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
                    padding: 0 62px;
                    background: white;
                }}
                
                /* Header Section */
                .header {{
                    margin-bottom: 24px;
                }}
                
                .title {{
                    font-size: 32px;
                    font-weight: 600;
                    color: #000000;
                    letter-spacing: -0.5px;
                    margin-bottom: 20px;
                }}
                
                .company-info {{
                    margin-top: 12px;
                }}
                
                .company-row {{
                    display: flex;
                    align-items: center;
                    margin-bottom: 8px;
                }}
                
                .company-icon {{
                    width: 14px;
                    height: 14px;
                    margin-right: 10px;
                    flex-shrink: 0;
                    filter: grayscale(100%) brightness(0.4);
                }}
                
                .company-label {{
                    font-size: 11px;
                    color: #000000;
                    font-weight: 700;
                    margin-right: 4px;
                }}
                
                .company-value {{
                    font-size: 11px;
                    color: #000000;
                    font-weight: 400;
                }}
                
                /* Divider */
                .divider {{
                    border-bottom: 1px solid #D3D3D3;
                    margin: 20px 0;
                }}
                
                /* Project Section */
                .project-section {{
                    display: flex;
                    align-items: center;
                    margin-bottom: 16px;
                }}
                
                .project-icon {{
                    width: 16px;
                    height: 16px;
                    margin-right: 8px;
                }}
                
                .project-name {{
                    font-size: 18px;
                    font-weight: 400;
                    color: #000000;
                }}
                
                /* Summary Box */
                .summary-box {{
                    background-color: rgba(28, 185, 126, 0.12);
                    border-radius: 4px;
                    padding: 16px;
                    margin-bottom: 20px;
                    display: inline-block;
                    min-width: 220px;
                }}
                
                .summary-row {{
                    display: flex;
                    justify-content: space-between;
                    margin-bottom: 4px;
                }}
                
                .summary-row:last-child {{
                    margin-bottom: 0;
                }}
                
                .summary-label {{
                    font-size: 10px;
                    font-weight: 700;
                    color: #000000;
                }}
                
                .summary-value {{
                    font-size: 10px;
                    color: #000000;
                    font-weight: 400;
                }}
                
                /* Profile Section */
                .profile-section {{
                    margin-bottom: 32px;
                    page-break-inside: avoid;
                }}
                
                .profile-title {{
                    font-size: 16px;
                    font-weight: 600;
                    color: #000000;
                    margin-bottom: 12px;
                }}
                
                /* Table */
                .parts-table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin-top: 9px;
                    page-break-inside: auto;
                }}
                
                .parts-table tr {{
                    page-break-inside: avoid;
                    page-break-after: auto;
                }}
                
                .parts-table thead {{
                    display: table-header-group;
                    border-bottom: 1px solid #D3D3D3;
                    background-color: #F3F4F6;
                }}
                
                .parts-table th {{
                    text-align: left;
                    padding: 6px 10px;
                    font-size: 8px;
                    font-weight: 700;
                    color: #A5A7A9;
                    text-transform: uppercase;
                }}
                
                .parts-table td {{
                    padding: 6px 10px;
                    font-size: 9px;
                    color: #000000;
                    border-bottom: 0.5px solid #F0F0F0;
                }}
                
                .parts-table th:nth-child(3),
                .parts-table th:nth-child(4),
                .parts-table th:nth-child(5),
                .parts-table th:nth-child(6),
                .parts-table td:nth-child(3),
                .parts-table td:nth-child(4),
                .parts-table td:nth-child(5),
                .parts-table td:nth-child(6) {{
                    text-align: right;
                }}
                
            </style>
        </head>
        <body>
            <!-- Header with Company Info -->
            <div class="header">
                <div class="title">Parts List</div>
                
                <div class="company-info">
{company_info_rows}
                </div>
            </div>
            
            <div class="divider"></div>
            
            <!-- Project Name -->
            <div class="project-section">
                <img src="data:image/svg+xml;base64,{icons.get('project', '')}" class="project-icon" />
                <span class="project-name">{project_name}</span>
            </div>
            
            <!-- Summary Box -->
            <div class="summary-box">
                <div class="summary-row">
                    <span class="summary-label">Total Parts:</span>
                    <span class="summary-value">{total_parts}</span>
                </div>
                <div class="summary-row">
                    <span class="summary-label">Total Weight:</span>
                    <span class="summary-value">{total_weight_kg:.2f} kg</span>
                </div>
                <div class="summary-row">
                    <span class="summary-label">Profiles:</span>
                    <span class="summary-value">{len([p for p in parts_by_profile.keys() if selected_parts.get(p)])}</span>
                </div>
            </div>
            
            <!-- Profile Sections -->
            {profile_sections_html}
            
        </body>
        </html>
        '''
        
        return html


class PartsListExcelGenerator:
    """Generates Parts Selection List Excel files using openpyxl."""

    def generate_excel(
        self,
        parts_by_profile: Dict[str, List[Dict]],
        selected_parts: Dict[str, List[str]],
        project_name: str,
        company_details: Dict[str, str]
    ) -> bytes:
        """Generate Parts List Excel from parts data."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime
        from io import BytesIO

        wb = Workbook()
        ws = wb.active
        ws.title = "Parts List"

        # Define styles
        header_fill = PatternFill(start_color="1CB97E", end_color="1CB97E", fill_type="solid")
        header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
        
        title_font = Font(name='Arial', size=16, bold=True)
        subtitle_font = Font(name='Arial', size=10, color="666666")
        
        profile_header_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        profile_header_font = Font(name='Arial', size=10, bold=True, color="6B7280")
        
        cell_font = Font(name='Arial', size=10)
        border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )

        current_row = 1

        # Title
        ws.merge_cells(f'A{current_row}:F{current_row}')
        title_cell = ws[f'A{current_row}']
        title_cell.value = "Parts List"
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        current_row += 2

        # Company details
        current_date = datetime.now().strftime("%d %B, %Y")
        
        company_info = [
            ('Company name:', company_details.get('companyName', 'N/A')),
            ('Address:', company_details.get('address', '')) if company_details.get('address') else None,
            ('Email:', company_details.get('email', '')) if company_details.get('email') else None,
            ('Phone:', company_details.get('phoneNumber', '')) if company_details.get('phoneNumber') else None,
            ('Date:', current_date)
        ]
        
        for info in company_info:
            if info:
                ws[f'A{current_row}'] = info[0]
                ws[f'A{current_row}'].font = Font(name='Arial', size=9, bold=True)
                ws[f'B{current_row}'] = info[1]
                ws[f'B{current_row}'].font = Font(name='Arial', size=9)
                current_row += 1
        
        current_row += 1

        # Project name
        ws[f'A{current_row}'] = "Project:"
        ws[f'A{current_row}'].font = Font(name='Arial', size=10, bold=True)
        ws[f'B{current_row}'] = project_name
        ws[f'B{current_row}'].font = Font(name='Arial', size=10)
        current_row += 2

        # Calculate totals
        total_parts = 0
        total_weight_kg = 0.0

        # Summary box
        ws[f'A{current_row}'] = "Total Parts:"
        ws[f'A{current_row}'].font = Font(name='Arial', size=9, bold=True)
        ws[f'A{current_row}'].fill = PatternFill(start_color="E8F5F0", end_color="E8F5F0", fill_type="solid")
        
        ws[f'A{current_row + 1}'] = "Total Weight:"
        ws[f'A{current_row + 1}'].font = Font(name='Arial', size=9, bold=True)
        ws[f'A{current_row + 1}'].fill = PatternFill(start_color="E8F5F0", end_color="E8F5F0", fill_type="solid")
        
        ws[f'A{current_row + 2}'] = "Profiles:"
        ws[f'A{current_row + 2}'].font = Font(name='Arial', size=9, bold=True)
        ws[f'A{current_row + 2}'].fill = PatternFill(start_color="E8F5F0", end_color="E8F5F0", fill_type="solid")

        # We'll update these after calculating
        summary_row = current_row
        current_row += 4

        # Generate profile sections
        for profile_name, parts_list in parts_by_profile.items():
            selected_part_numbers = set(selected_parts.get(profile_name, []))
            
            # Filter only selected parts
            selected_parts_list = [
                part for part in parts_list 
                if str(part.get('part_number', '')) in selected_part_numbers or 
                   str(part.get('product_id', '')) in selected_part_numbers
            ]
            
            if not selected_parts_list:
                continue
            
            # Sort by part number
            selected_parts_list.sort(key=lambda p: str(p.get('part_number', '')))
            
            # Calculate profile totals
            profile_total_weight = sum(
                part.get('weight', 0) * part.get('quantity', 1) 
                for part in selected_parts_list
            )
            total_parts += len(selected_parts_list)
            total_weight_kg += profile_total_weight

            # Profile title
            ws.merge_cells(f'A{current_row}:F{current_row}')
            profile_title_cell = ws[f'A{current_row}']
            profile_title_cell.value = f"{profile_name} • {len(selected_parts_list)} parts"
            profile_title_cell.font = Font(name='Arial', size=12, bold=True)
            profile_title_cell.alignment = Alignment(horizontal='left', vertical='center')
            current_row += 1

            # Table headers
            headers = ['Part Number', 'Part Name', 'Length (mm)', 'Quantity', 'Weight (kg)', 'Total Weight (kg)']
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = header
                cell.font = profile_header_font
                cell.fill = profile_header_fill
                cell.alignment = Alignment(horizontal='center' if col_idx > 2 else 'left', vertical='center')
                cell.border = border
            
            current_row += 1

            # Table data
            for part in selected_parts_list:
                part_weight = part.get('weight', 0)
                quantity = part.get('quantity', 1)
                part_total_weight = part_weight * quantity

                row_data = [
                    part.get('part_number', 'Unknown'),
                    part.get('element_name', 'Unnamed'),
                    part.get('length', 0),
                    quantity,
                    round(part_weight, 2),
                    round(part_total_weight, 2)
                ]

                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.value = value
                    cell.font = cell_font
                    cell.alignment = Alignment(horizontal='right' if col_idx > 2 else 'left', vertical='center')
                    cell.border = border
                
                current_row += 1
            
            current_row += 2  # Space between profiles

        # Update summary values
        ws[f'B{summary_row}'] = total_parts
        ws[f'B{summary_row}'].font = Font(name='Arial', size=9)
        ws[f'B{summary_row}'].fill = PatternFill(start_color="E8F5F0", end_color="E8F5F0", fill_type="solid")
        
        ws[f'B{summary_row + 1}'] = f"{total_weight_kg:.2f} kg"
        ws[f'B{summary_row + 1}'].font = Font(name='Arial', size=9)
        ws[f'B{summary_row + 1}'].fill = PatternFill(start_color="E8F5F0", end_color="E8F5F0", fill_type="solid")
        
        ws[f'B{summary_row + 2}'] = len([p for p in parts_by_profile.keys() if selected_parts.get(p)])
        ws[f'B{summary_row + 2}'].font = Font(name='Arial', size=9)
        ws[f'B{summary_row + 2}'].fill = PatternFill(start_color="E8F5F0", end_color="E8F5F0", fill_type="solid")

        # Set column widths
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 18

        # Save to bytes
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer.getvalue()
